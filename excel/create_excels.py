# excel.create_excels.py
from config.settings import WENCO_EXCEL_REPORT_NAME
from datetime import datetime, timedelta
from database.db_helpers import get_db_connection
from excel_helpers import query_sales_summaries, query_cars_summaries, query_gp_summaries, query_labor_revenue_summaries, query_gs_hours, query_tech_hours, query_timesheet_discrepancies, query_tech_flagged_hours, query_bigo_sales_details
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def fetch_shops_sales(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo_sales = query_sales_summaries("bigo", end_date, start_date)
        midas_sales = query_sales_summaries("midas", end_date, start_date)

        # Combine both results into a single list
        all_sales = bigo_sales + midas_sales

        sales_data_sorted = sorted(all_sales, key=lambda x: x[0])

        return sales_data_sorted
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_shops_cars(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo_cars = query_cars_summaries("bigo", end_date, start_date)
        midas_cars = query_cars_summaries("midas", end_date, start_date)

        # Combine both results into a single list
        all_cars = bigo_cars + midas_cars
        cars_data_sorted = sorted(all_cars, key=lambda x: x[0])

        return cars_data_sorted
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_shops_gp(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_gp_summaries("bigo", end_date, start_date)
        midas = query_gp_summaries("midas", end_date, start_date)

        # Combine both results into a single list
        all_data = bigo + midas
        sorted_data = sorted(all_data, key=lambda x: x[0])

        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_shops_labor_revenue(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_labor_revenue_summaries("bigo", end_date, start_date)
        midas = query_labor_revenue_summaries("midas", end_date, start_date)

        # Combine both results into a single list
        all_data = bigo + midas
        sorted_data = sorted(all_data, key=lambda x: x[0])

        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_gs_hours_worked(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_gs_hours("bigo", end_date, start_date)
        midas = query_gs_hours("midas", end_date, start_date)

        # Combine both results into a single list
        all_data = bigo + midas
        sorted_data = sorted(all_data, key=lambda x: x[0])

        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_tech_hours_worked(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_tech_hours("bigo", end_date, start_date)
        midas = query_tech_hours("midas", end_date, start_date)

        # Combine both results into a single list
        all_data = bigo + midas
        sorted_data = sorted(all_data, key=lambda x: x[0])

        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_timesheet_discrepancies(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_timesheet_discrepancies("bigo", end_date, start_date)
        midas = query_timesheet_discrepancies("midas", end_date, start_date)

        # Combine both results into a single list
        all_discrepancies = {"bigo": bigo, "midas": midas}

        return all_discrepancies
    except Exception as e:
        print(f"Exception {e} occurred")
        return []


def fetch_tech_flagged_hours(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo = query_tech_flagged_hours("bigo", end_date, start_date)
        midas = query_tech_flagged_hours("midas", end_date, start_date)

        # Combine both results into a single list
        all_data = bigo + midas
        sorted_data = sorted(all_data, key=lambda x: x[0])

        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []

def fetch_bigo_specific_details(start_date):
    # Calculate the end date (7 days prior)
    end_date = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        # Ensure the database connection can be established
        conn = get_db_connection()
        conn.close()  # Close the connection immediately after the test

        # Fetch sales summaries for both Big O Tires and Midas
        bigo_data = query_bigo_sales_details(end_date, start_date)
        sorted_data = sorted(bigo_data, key=lambda x: x[0])


        return sorted_data
    except Exception as e:
        print(f"Exception {e} occurred")
        return []


def create_excel_report():
    filename = WENCO_EXCEL_REPORT_NAME
    start_date = input("Enter a date (YYYY-MM-DD) to retrieve the prior 7 days data and export to excel: ")

    # Fetch the data
    sales_data = fetch_shops_sales(start_date)
    cars_data = fetch_shops_cars(start_date)
    gp_data = fetch_shops_gp(start_date)
    lab_rev_data = fetch_shops_labor_revenue(start_date)
    gs_hours_worked = fetch_gs_hours_worked(start_date)
    tech_hours_worked = fetch_tech_hours_worked(start_date)
    all_discrepancies = fetch_timesheet_discrepancies(start_date)
    tech_flagged_hours = fetch_tech_flagged_hours(start_date)
    bigo_specific_data = fetch_bigo_specific_details(start_date)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Define a helper function to create a worksheet
    def create_sheet(title, headers, data):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in data:
            ws.append(row)
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Create sheets with data
    create_sheet("Sales Data", ["Store Number", "Total Revenue w/ Supplies"], sales_data)
    create_sheet("Car Count Data", ["Store Number", "Total Car Count"], cars_data)
    create_sheet("Shops GP", ["Store Number", "Total Gross Profit"], gp_data)
    create_sheet("Shops Labor Revenue", ["Store Number", "Total Labor Revenue"], lab_rev_data)
    create_sheet("GS Hours Worked", ["Store Number", "Total GS Hours"], gs_hours_worked)
    create_sheet("Tech Hours Worked", ["Store Number", "Total Tech Hours"], tech_hours_worked)
    create_sheet("Tech Hours Flagged", ["Store Number", "Total Tech Hours Flagged"], tech_flagged_hours)
    create_sheet("Big O Sales Details", ["Store Number", "Total Alignments", "Total Tires", "Total Nitrogens"], bigo_specific_data)

    # Create a sheet for Timesheet Discrepancies
    ws = wb.create_sheet("Timesheet Discrepancies")
    ws.append(["Store Type", "First Name", "Last Name"])
    for store, disc in all_discrepancies.items():
        for first_name, last_name in disc:
            ws.append([store.capitalize(), first_name, last_name])
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Remove the default sheet created by openpyxl
    del wb['Sheet']

    # Save the workbook
    wb.save(filename)
    print(f"Report saved as {filename}")

def print_wenco_fetch():
    start_date = input("Enter a date (YYYY-MM-DD) to retrieve the prior 7 days data and export to excel: ")

    # Fetch the sales data
    sales_data = fetch_shops_sales(start_date)
    cars_data = fetch_shops_cars(start_date)
    gp_data = fetch_shops_gp(start_date)
    lab_rev_data = fetch_shops_labor_revenue(start_date)
    gs_hours_worked = fetch_gs_hours_worked(start_date)
    tech_hours_worked = fetch_tech_hours_worked(start_date)
    all_discrepancies = fetch_timesheet_discrepancies(start_date)
    tech_flagged_hours = fetch_tech_flagged_hours(start_date)
    bigo_specific_data = fetch_bigo_specific_details(start_date)

    # Print the fetched data in a listed format
    print("\nSales Data:")
    for store_number, total_revenue in sales_data:
        print(f"Store Number: {store_number}, Total Revenue w/ Supplies: ${total_revenue:.2f}")

    print("\nCar Count Data:")
    for store_number, total_car_count in cars_data:
        print(f"Store Number: {store_number}, Total Car Count: {total_car_count}")

    print("\nShops GP:")
    for store_number, total_gp in gp_data:
        print(f"Store Number: {store_number}, Total Gross Profit: {total_gp}")

    print("\nShops Labor Revenue:")
    for store_number, total_lab_rev in lab_rev_data:
        print(f"Store Number: {store_number}, Total Labor Revenue: {total_lab_rev}")

    print("\nGS HOURS WORKED:")
    for store_number, gs_hours in gs_hours_worked:
        print(f"Store Number: {store_number}, Total Labor Revenue: {gs_hours}")

    print("\nTech HOURS WORKED:")
    for store_number, tech_hours in tech_hours_worked:
        print(f"Store Number: {store_number}, Total Labor Revenue: {tech_hours}")

    print("\nTimesheet Discrepancies:")
    for store, disc in all_discrepancies.items():
        print(f"\nStore Type: {store.capitalize()}")
        for first_name, last_name in disc:
            print(f"First Name: {first_name}, Last Name: {last_name}")

    print("\nTech Hours Flagged (Midas, Big O Tires):")
    for store_number, total_tech_hours in tech_flagged_hours:
        print(f"Store Number: {store_number}, Total Tech Hours Flagged: {total_tech_hours:.2f}")

    print("\nBig O Tires Sales Details (Alignments, Tires, Nitrogens):")
    for store_number, total_alignments, total_tires, total_nitrogens in bigo_specific_data:
        print(
            f"Store Number: {store_number}, Total Alignments: {total_alignments}, Total Tires: {total_tires}, Total Nitrogens: {total_nitrogens}")

if __name__ == "__main__":
    create_excel_report()

